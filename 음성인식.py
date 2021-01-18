import time

import selenium
from selenium import webdriver
from selenium.webdriver import ActionChains

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait

from openpyxl import load_workbook

#엑셀 불러오기
load_wb = load_workbook("moves.xlsx", data_only=True)
movetable = load_wb['Sheet1']
movename = []
movehex = []

for i in range(1,270):
	movename.append(movetable.cell(i,4).value)
	movehex.append(movetable.cell(i,1).value)

# 주소 설정
URL = 'https://speech-api.kakao.com/#none'
driver = webdriver.Chrome(executable_path='chromedriver')
time.sleep(0.5)
driver.get(url=URL)




#시작 대기
startinput = input("시작할 때 아무 키나 누르세요")

link_voice = driver.find_element_by_class_name('link_voice')
txt_result = driver.find_element_by_class_name('txt_result')

print(txt_result.text)
print(link_voice.get_attribute("class"))

while True:
	current_link = link_voice.get_attribute("class")

	if current_link == 'link_voice on' :
		while True : 
			current_link = link_voice.get_attribute("class")
			if current_link == 'link_voice' :
				current_text = txt_result.text
				current_text = current_text.replace(" ","")
				if current_text in movename:
					voice2move = movename.index(current_text)
				else :
					voice2move = 122
				print(movehex[voice2move])
				selection = movehex[voice2move]


				#기술입력 메모장
				textfile = open("movefile.lua", 'w')
				final = "a = 1\nhex = %s" % selection
				textfile.write(final)
				textfile.close()
				time.sleep(0.5)
				newfile = open("movefile.lua", 'w')
				newfile.write(final)
				final = "\na = 0\nhex = %s" % selection
				newfile.write(final)
				newfile.close()
				break
			time.sleep(0.5)


	time.sleep(1)
print(txt_result.text)


